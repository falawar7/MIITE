"""
MIITE Off the Record — Booking App (Streamlit)

Click any slot in the schedule to open a small popup:
  - Free slot  -> enter company name -> book it
  - Booked slot -> see who booked it -> cancel if needed

Run locally:   streamlit run app.py
Deploy:        push to GitHub -> https://streamlit.io/cloud (1-click)
"""

from datetime import datetime, time, timedelta
from pathlib import Path

import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, PatternFill

# ---- Configuration ----
EXCEL_FILE = "schedule.xlsx"
SHEET_NAME = "Schedule"
DAYS = ["MON", "TUE", "WED", "THU", "FRI"]
WEEK_START = datetime(2026, 5, 4)
HOURS = list(range(10, 18))
DATES = [(WEEK_START + timedelta(days=i)).strftime("%Y-%m-%d") for i in range(len(DAYS))]
# Friendly format used in the UI only (e.g. 'May 4')
# Note: '%-d' isn't portable to Windows, so we build it manually
DATES_DISPLAY = [
    (WEEK_START + timedelta(days=i)).strftime("%b ")
    + str((WEEK_START + timedelta(days=i)).day)
    for i in range(len(DAYS))
]

# ---- Branding ----
LOGO_LEFT = "logo_left.png"
LOGO_RIGHT = "logo_right.png"
HEADER_TITLE = "MIITE Off the Record"
HEADER_SUBTITLE = "Week of May 4–8, 2026 · 1-hour slots, 10:00–17:00"


# ---- Excel data layer (unchanged from before) ----
def init_excel(path: str = EXCEL_FILE) -> None:
    if Path(path).exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws["D6"] = "Schedule Sheet"
    ws["G6"] = "TIME INTERVAL"
    ws["D7"] = time(9, 0)
    ws["G7"] = 60
    ws["H7"] = "(In Minutes)"
    ws["B9"] = "TIME"
    for i, d in enumerate(DAYS):
        ws.cell(row=9, column=4 + i, value=d)
        ws.cell(row=10, column=4 + i, value=DATES[i])
    for h_idx, h in enumerate(HOURS):
        excel_row = 11 + h_idx * 2
        ws.cell(row=excel_row, column=1, value=time(h, 0))
        for d_idx in range(len(DAYS)):
            ws.cell(row=excel_row, column=4 + d_idx, value="No")
    wb.save(path)


def load_bookings(path: str = EXCEL_FILE) -> dict:
    init_excel(path)
    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    bookings = {}
    for h_idx, h in enumerate(HOURS):
        excel_row = 11 + h_idx * 2
        for d_idx in range(len(DAYS)):
            cell = ws.cell(row=excel_row, column=4 + d_idx).value
            if cell is None:
                continue
            txt = str(cell).strip()
            if txt.lower().startswith("yes"):
                company = txt.split("—", 1)[-1].strip() if "—" in txt else ""
                bookings[(d_idx, h)] = company or "(unnamed)"
    return bookings


def save_booking(d_idx: int, hour: int, company, path: str = EXCEL_FILE) -> None:
    init_excel(path)
    wb = load_workbook(path)
    ws = wb[SHEET_NAME]
    h_idx = HOURS.index(hour)
    excel_row = 11 + h_idx * 2
    cell = ws.cell(row=excel_row, column=4 + d_idx)
    if company:
        cell.value = f"Yes — {company}"
        cell.fill = PatternFill("solid", start_color="FAECE7")
    else:
        cell.value = "No"
        cell.fill = PatternFill("solid", start_color="EAF3DE")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    wb.save(path)


# ---- Page setup ----
st.set_page_config(
    page_title="MIITE Off the Record · Booking",
    page_icon="📅",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ---- Custom styling ----
st.markdown(
    """
    <style>
    /* Tighten up Streamlit's default padding */
    .block-container { padding-top: 1.2rem; padding-bottom: 2rem; max-width: 1100px; }

    @media (max-width: 600px) {
      .block-container { padding-left: 0.5rem !important; padding-right: 0.5rem !important;
                         padding-top: 0.5rem !important; }
    }

    /* Slot buttons fill their column */
    .stButton > button { width: 100%; padding: 10px 4px; font-size: 12px;
                         font-weight: 500; border-radius: 6px;
                         min-height: 44px; transition: transform 0.08s; }
    .stButton > button:hover { transform: translateY(-1px); }
    .stButton > button:active { transform: scale(0.97); }

    /* Free vs booked slot colors */
    .stButton > button[kind="secondary"] {
        background: #EAF3DE !important; border: 1px solid #97C459 !important;
        color: #173404 !important; }
    .stButton > button[kind="primary"] {
        background: #FAECE7 !important; border: 1px solid #D85A30 !important;
        color: #4A1B0C !important; }

    /* === FIXED CELL HEIGHT for slot buttons inside the grid ===
       Prevents long names from making a row taller and pushing other cells
       out of alignment. Anything that doesn't fit is hidden with an ellipsis.
       Footer buttons are excluded by their st-key class. */
    div[data-testid="stColumn"] .stButton > button {
        height: 44px !important;
        min-height: 44px !important;
        max-height: 44px !important;
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
        line-height: 1 !important;
    }
    /* Inner span Streamlit renders for button text — also clamp its overflow */
    div[data-testid="stColumn"] .stButton > button > div,
    div[data-testid="stColumn"] .stButton > button p {
        white-space: nowrap !important;
        overflow: hidden !important;
        text-overflow: ellipsis !important;
        max-width: 100% !important;
        margin: 0 !important;
    }
    /* Footer buttons can have their own height/wrapping rules below */

    /* Footer action buttons — neutral grey style, matched height.
       Targeted by the button keys (footer_refresh, footer_download) which
       Streamlit emits as 'st-key-<key>' classes on the wrapping element. */
    .st-key-footer_refresh button,
    .st-key-footer_download button,
    div[data-testid="stDownloadButton"] button {
        background: #ffffff !important;
        border: 1px solid #d4d4d8 !important;
        color: #18181b !important;
        font-weight: 500 !important;
        min-height: 44px !important;
        padding: 8px 16px !important;
        border-radius: 8px !important;
        width: 100% !important;
        transition: all 0.15s !important;
        transform: none !important;
    }
    .st-key-footer_refresh button:hover,
    .st-key-footer_download button:hover,
    div[data-testid="stDownloadButton"] button:hover {
        background: #f4f4f5 !important;
        border-color: #a1a1aa !important;
        transform: none !important;
    }

    /* Time-column label */
    .miite-time { padding: 12px 4px; text-align: center; color: #666;
                  font-variant-numeric: tabular-nums; font-size: 13px; font-weight: 500; }

    /* Day-header label */
    .miite-day { text-align: center; font-weight: 500; padding: 6px 2px;
                 font-size: 13px; }
    .miite-day .miite-date { color: #999; font-weight: 400; font-size: 11px;
                             display: block; }

    /* Stats row */
    .miite-stats { display: flex; gap: 8px; margin: 4px 0 16px; flex-wrap: wrap; }
    .miite-stat { background: #f5f5f4; padding: 8px 14px; border-radius: 8px;
                  flex: 1 1 auto; min-width: 90px; text-align: center; }
    .miite-stat .lbl { font-size: 11px; color: #666; }
    .miite-stat .val { font-size: 18px; font-weight: 600; }

    /* Header: title on top, logos side-by-side underneath */
    .miite-header { padding: 8px 0 16px; border-bottom: 1px solid #eee;
                    margin-bottom: 18px; text-align: center; }
    .miite-header h1 { margin: 0; font-size: 24px; line-height: 1.2; }
    .miite-header .miite-sub { color: #666; font-size: 13px; margin-top: 4px; }
    .miite-header .miite-logos { display: flex; justify-content: center;
                                 align-items: center; gap: 28px; margin-top: 14px;
                                 flex-wrap: wrap; }
    .miite-header img.miite-logo-left  { height: 90px; max-width: 240px; object-fit: contain; }
    .miite-header img.miite-logo-right { height: 56px; max-width: 160px; object-fit: contain; }

    /* === KEY FIX FOR MOBILE ===
       Force Streamlit's columns to stay horizontal at any screen width.
       Without this, columns stack vertically on phones, breaking the grid. */
    div[data-testid="stHorizontalBlock"] {
        flex-wrap: nowrap !important;
        gap: 4px !important;
    }
    div[data-testid="stHorizontalBlock"] > div[data-testid="stColumn"] {
        min-width: 0 !important;
        flex-shrink: 1 !important;
    }
    /* Make buttons inside narrow columns stay readable */
    div[data-testid="stColumn"] .stButton > button {
        padding: 8px 2px !important;
        font-size: 11px !important;
        min-height: 38px !important;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
    }

    @media (max-width: 600px) {
      .miite-header img.miite-logo-left  { height: 60px !important; max-width: 160px !important; }
      .miite-header img.miite-logo-right { height: 38px !important; max-width: 110px !important; }
      .miite-header h1 { font-size: 18px; }
      .miite-header .miite-sub { font-size: 12px; }
      .miite-header .miite-logos { gap: 18px; margin-top: 10px; }
      div[data-testid="stColumn"] .stButton > button {
        height: 38px !important; min-height: 38px !important; max-height: 38px !important;
        padding: 4px 1px !important; font-size: 9px !important; }
      .miite-time { font-size: 10px; padding: 8px 1px; }
      .miite-day { font-size: 10px; }
      .miite-day .miite-date { font-size: 9px; }
      .miite-stat { padding: 6px 8px; min-width: 70px; }
      .miite-stat .val { font-size: 15px; }
      .miite-stat .lbl { font-size: 10px; }
      div[data-testid="stHorizontalBlock"] { gap: 2px !important; }
    }

    @media (max-width: 400px) {
      div[data-testid="stColumn"] .stButton > button {
        height: 34px !important; min-height: 34px !important; max-height: 34px !important;
        font-size: 8px !important; padding: 2px 1px !important; }
      .miite-time { font-size: 9px; }
      .miite-day { font-size: 9px; }
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# ---- Header with logos ----
def render_header():
    def logo_html(path: str, css_class: str) -> str:
        p = Path(path)
        if not p.exists():
            return ""
        import base64, mimetypes
        mime = mimetypes.guess_type(p.name)[0] or "image/png"
        data = base64.b64encode(p.read_bytes()).decode()
        return f'<img class="{css_class}" src="data:{mime};base64,{data}" alt="logo" />'

    left = logo_html(LOGO_LEFT, "miite-logo-left")
    right = logo_html(LOGO_RIGHT, "miite-logo-right")
    logos = (left + right) if (left or right) else ""

    st.markdown(
        f"""
        <div class="miite-header">
            <h1>{HEADER_TITLE}</h1>
            <div class="miite-sub">{HEADER_SUBTITLE}</div>
            {f'<div class="miite-logos">{logos}</div>' if logos else ''}
        </div>
        """,
        unsafe_allow_html=True,
    )


# ---- Popup dialog ----
@st.dialog("Slot details")
def slot_dialog(d_idx: int, hour: int):
    """A small popup that opens when a slot is clicked."""
    bookings = load_bookings()
    booking = bookings.get((d_idx, hour))

    label = f"**{DAYS[d_idx]} · {DATES_DISPLAY[d_idx]} · {hour:02d}:00 – {hour + 1:02d}:00**"
    st.markdown(label)

    if booking:
        # Booked → show details + cancel option
        st.info(f"Currently booked by **{booking}**")
        col_a, col_b = st.columns(2)
        if col_a.button("🗑️ Cancel booking", type="primary", use_container_width=True):
            save_booking(d_idx, hour, None)
            st.session_state["last_msg"] = (
                f"🗑️ Cancelled **{booking}** ({DAYS[d_idx]} {hour:02d}:00)"
            )
            st.rerun()
        if col_b.button("Keep it", use_container_width=True):
            st.rerun()
    else:
        # Free → ask for company name
        st.success("This slot is free — book it below:")
        with st.form("book_form", clear_on_submit=False):
            company = st.text_input(
                "Company name",
                placeholder="e.g. Acme Corp",
                key=f"company_{d_idx}_{hour}",
            )
            col_a, col_b = st.columns(2)
            confirm = col_a.form_submit_button(
                "✅ Confirm booking", type="primary", use_container_width=True
            )
            close = col_b.form_submit_button("Cancel", use_container_width=True)

            if confirm:
                if not company.strip():
                    st.error("Please enter a company name.")
                else:
                    save_booking(d_idx, hour, company.strip())
                    st.session_state["last_msg"] = (
                        f"✅ Booked **{DAYS[d_idx]} {hour:02d}:00** for **{company.strip()}**"
                    )
                    st.rerun()
            elif close:
                st.rerun()


# ---- Main rendering ----
render_header()

# Status message from previous action (rendered just under the header)
last_msg = st.session_state.pop("last_msg", None)
if last_msg:
    st.success(last_msg)

bookings = load_bookings()
total = len(DAYS) * len(HOURS)
booked = len(bookings)

# Stats row
st.markdown(
    f"""
    <div class="miite-stats">
      <div class="miite-stat"><div class="lbl">Total</div><div class="val">{total}</div></div>
      <div class="miite-stat"><div class="lbl">Booked</div>
          <div class="val" style="color:#993C1D">{booked}</div></div>
      <div class="miite-stat"><div class="lbl">Free</div>
          <div class="val" style="color:#0F6E56">{total - booked}</div></div>
      <div class="miite-stat"><div class="lbl">Utilization</div>
          <div class="val">{round(booked / total * 100)}%</div></div>
    </div>
    """,
    unsafe_allow_html=True,
)

st.caption("👆 Tap any slot to book it or to cancel an existing booking.")

# ---- Schedule grid ----
# Layout: time column + 7 day columns. Streamlit columns produce a tap-friendly grid.
cols = st.columns([1] + [2] * len(DAYS))

# Header row
cols[0].markdown('<div class="miite-day">&nbsp;</div>', unsafe_allow_html=True)
for i, d in enumerate(DAYS):
    cols[i + 1].markdown(
        f'<div class="miite-day">{d}<span class="miite-date">{DATES_DISPLAY[i]}</span></div>',
        unsafe_allow_html=True,
    )

# Body rows
for h in HOURS:
    row = st.columns([1] + [2] * len(DAYS))
    row[0].markdown(f'<div class="miite-time">{h:02d}:00</div>', unsafe_allow_html=True)
    for d_idx in range(len(DAYS)):
        company = bookings.get((d_idx, h))
        if company:
            # Cell is narrow on mobile — keep label short, full name shows in popup
            label = company if len(company) <= 6 else company[:5] + "…"
            btype = "primary"  # booked
        else:
            label = "·"
            btype = "secondary"  # free
        if row[d_idx + 1].button(
            label,
            key=f"slot_{d_idx}_{h}",
            type=btype,
            use_container_width=True,
            help=company if company else "Free — tap to book",
        ):
            slot_dialog(d_idx, h)

# ---- Footer actions ----
st.divider()
fcol1, fcol2 = st.columns(2, gap="small")
with fcol1:
    if st.button("🔄 Refresh", use_container_width=True, key="footer_refresh"):
        st.rerun()
with fcol2:
    if Path(EXCEL_FILE).exists():
        with open(EXCEL_FILE, "rb") as f:
            st.download_button(
                "⬇ Download .xlsx",
                data=f.read(),
                file_name="MIITE_Schedule.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="footer_download",
            )

# ---- Developer credit ----
st.markdown(
    """
    <div style="text-align: center; margin-top: 24px; padding: 14px 0 8px;
                color: #888; font-size: 12px; border-top: 1px solid #eee;">
        App developed by <b>Faisal Elawar</b>
    </div>
    """,
    unsafe_allow_html=True,
)